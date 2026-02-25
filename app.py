from flask import Flask, request, send_file, Response
import pdfplumber, re, io, json, urllib.parse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Base directory of the app (where app.py lives) — works on Windows, Linux, macOS
BASE_DIR = Path(__file__).resolve().parent

app = Flask(__name__)

@app.route("/")
def index():
    index_path = BASE_DIR / "index.html"
    return open(index_path, encoding="utf-8").read(), 200, {"Content-Type": "text/html"}

@app.route("/analyze", methods=["POST"])
def analyze():
    if "pdf" not in request.files:
        return {"error": "No file"}, 400
    
    pdf_bytes = request.files["pdf"].read()
    months = ['10-2025','11-2025','12-2025','01-2026']
    
    # ── Parse PDF ──────────────────────────────────────────────
    clients = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text: continue
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            client_name = client_code = None
            for line in lines[:8]:
                m = re.match(r'^(.+?)\s+\(([A-Z0-9]+)\)\s*$', line)
                if m and not any(x in m.group(1) for x in ['TRU Funding','Client Analysis','As Of','Printed']):
                    if 2 <= len(m.group(2)) <= 10:
                        client_name = m.group(1).strip()
                        client_code = m.group(2).strip()
                        break
            if not client_name: continue
            sales = {}
            for month in months:
                m = re.search(month.replace('-', r'\-') + r'\s+([\d,]+\.\d+)', text)
                sales[month] = float(m.group(1).replace(',','')) if m else 0.0
            clients.append({'name': client_name, 'code': client_code, 'sales': sales})

    # ── Stats ──────────────────────────────────────────────────
    jan_total = sum(c['sales'].get('01-2026', 0) for c in clients)
    up_count = down_count = 0
    for c in clients:
        vals = [c['sales'].get(m, 0) for m in months]
        jan = vals[3]; prev = vals[2] if vals[2]>0 else (vals[1] if vals[1]>0 else vals[0])
        if prev > 0:
            if jan > prev: up_count += 1
            elif jan < prev: down_count += 1

    # ── Preview top 5 by avg ───────────────────────────────────
    sorted_c = sorted(clients, key=lambda c: sum(c['sales'].get(m,0) for m in months)/4, reverse=True)
    preview = []
    for c in sorted_c[:5]:
        vals = [c['sales'].get(m,0) for m in months]
        avg = sum(vals)/4; jan = vals[3]
        prev = vals[2] if vals[2]>0 else (vals[1] if vals[1]>0 else vals[0])
        trend = (jan-prev)/prev if prev>0 else 0
        dir_label = '► FLAT'
        if prev>0 and jan>prev: dir_label = '▲ UP'
        elif prev>0 and jan<prev: dir_label = '▼ DOWN'
        preview.append({'name':c['name'],'code':c['code'],
            'oct':vals[0],'nov':vals[1],'dec':vals[2],'jan':vals[3],
            'avg':avg,'trend':trend,'dir':dir_label})

    # ── Build Excel ────────────────────────────────────────────
    wb = build_excel(clients)
    out = io.BytesIO(); wb.save(out); out.seek(0)

    resp = Response(out.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers['Content-Disposition'] = 'attachment; filename="TRU_Client_Sales_Analysis.xlsx"'
    resp.headers['X-Client-Count'] = str(len(clients))
    resp.headers['X-Jan-Total']    = str(jan_total)
    resp.headers['X-Up-Count']     = str(up_count)
    resp.headers['X-Down-Count']   = str(down_count)
    resp.headers['X-Preview']      = urllib.parse.quote(json.dumps(preview))
    resp.headers['Access-Control-Expose-Headers'] = 'X-Client-Count,X-Jan-Total,X-Up-Count,X-Down-Count,X-Preview'
    return resp


def build_excel(clients):
    months = ['10-2025','11-2025','12-2025','01-2026']
    wb = Workbook(); ws = wb.active; ws.title = "Client Sales Analysis"

    NAVY="1F4E79"; BLUE="2E75B6"; LB="D6E4F0"; LG="E2EFDA"
    LGR="F5F9FF"; WHITE="FFFFFF"; UP_F="E2EFDA"; DN_F="FCE4D6"; FL_F="FFF2CC"
    UP_C="375623"; DN_C="7B0000"; FL_C="7F6000"
    thin = Side(style='thin', color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    money = '$#,##0.00'

    # Title
    ws.merge_cells("A1:K1")
    c=ws["A1"]; c.value="TRU Funding LLC — Client Sales Analysis"
    c.font=Font(name="Arial",bold=True,size=14,color=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center')
    c.fill=PatternFill("solid",start_color="EBF3FB")
    ws.row_dimensions[1].height=32

    ws.merge_cells("A2:K2")
    c=ws["A2"]; c.value="As of October 31, 2025  ·  Printed: February 23, 2026"
    c.font=Font(name="Arial",italic=True,size=9,color="666666")
    c.alignment=Alignment(horizontal='center',vertical='center')
    c.fill=PatternFill("solid",start_color="F7FBFF")
    ws.row_dimensions[2].height=16
    ws.row_dimensions[3].height=8

    hdrs=["#","Client Name","Code","Oct-2025","Nov-2025","Dec-2025","Jan-2026","Average Sales","Last Month","Trend %","Direction"]
    for col,h in enumerate(hdrs,1):
        c=ws.cell(row=4,column=col,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        c.fill=PatternFill("solid",start_color=NAVY)
        c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        c.border=bdr
    ws.row_dimensions[4].height=28

    for i,w in enumerate([5,32,10,13,13,13,13,15,14,12,12],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    for i,client in enumerate(clients):
        row=i+5; alt=LGR if i%2==0 else WHITE
        s=client['sales']; vals=[s.get(m,0) for m in months]
        jan=vals[3]; prev=vals[2] if vals[2]>0 else (vals[1] if vals[1]>0 else vals[0])
        if prev==0: df,dc,arrow=FL_F,FL_C,"► FLAT"
        elif jan>prev: df,dc,arrow=UP_F,UP_C,"▲ UP"
        else: df,dc,arrow=DN_F,DN_C,"▼ DOWN"
        trend=(jan-prev)/prev if prev>0 else 0

        def sc(col,val,bg,bold=False,color="000000",fmt=None,ha='left'):
            c=ws.cell(row=row,column=col,value=val)
            c.font=Font(name="Arial",size=10,bold=bold,color=color)
            c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal=ha,vertical='center')
            c.border=bdr
            if fmt: c.number_format=fmt

        sc(1,i+1,alt,ha='center'); sc(2,client['name'],alt); sc(3,client['code'],alt,ha='center')
        for ci,v in enumerate(vals,4): sc(ci,v,alt,fmt=money,ha='right')
        c=ws.cell(row=row,column=8,value=f"=AVERAGE(D{row}:G{row})")
        c.font=Font(name="Arial",size=10); c.fill=PatternFill("solid",start_color=LB)
        c.alignment=Alignment(horizontal='right',vertical='center'); c.border=bdr; c.number_format=money
        c=ws.cell(row=row,column=9,value=f"=G{row}")
        c.font=Font(name="Arial",size=10); c.fill=PatternFill("solid",start_color=LG)
        c.alignment=Alignment(horizontal='right',vertical='center'); c.border=bdr; c.number_format=money
        sc(10,trend,df,bold=True,color=dc,fmt='+0.0%;-0.0%;0.0%',ha='center')
        sc(11,arrow,df,bold=True,color=dc,ha='center')
        ws.row_dimensions[row].height=18

    tr=len(clients)+5
    ws.merge_cells(f"A{tr}:C{tr}")
    c=ws.cell(row=tr,column=1,value="TOTALS")
    c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    c.fill=PatternFill("solid",start_color=BLUE); c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
    for ci in range(4,12):
        cl=get_column_letter(ci)
        val=f"=SUM({cl}5:{cl}{tr-1})" if ci<=9 else ""
        c=ws.cell(row=tr,column=ci,value=val)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        c.fill=PatternFill("solid",start_color=BLUE)
        c.alignment=Alignment(horizontal='right',vertical='center'); c.border=bdr
        if ci<=9: c.number_format=money
    ws.row_dimensions[tr].height=22
    ws.freeze_panes="D5"

    # Sheet 2
    ws2=wb.create_sheet("Trend Summary")
    ws2.merge_cells("A1:G1")
    c=ws2["A1"]; c.value="TRU Funding LLC — Sales Trend Summary"
    c.font=Font(name="Arial",bold=True,size=13,color=NAVY)
    c.alignment=Alignment(horizontal='center',vertical='center')
    c.fill=PatternFill("solid",start_color="EBF3FB"); ws2.row_dimensions[1].height=26

    for col,h in enumerate(["Rank","Client","Code","Average Sales","Jan-2026 Sales","vs Prev Month","Direction"],1):
        c=ws2.cell(row=2,column=col,value=h)
        c.font=Font(name="Arial",bold=True,color="FFFFFF",size=10)
        c.fill=PatternFill("solid",start_color=NAVY)
        c.alignment=Alignment(horizontal='center',vertical='center'); c.border=bdr
    ws2.row_dimensions[2].height=22

    for w,col in zip([6,35,10,16,16,14,12],"ABCDEFG"):
        ws2.column_dimensions[col].width=w

    sorted_c=sorted(clients,key=lambda c:sum(c['sales'].get(m,0) for m in months)/4,reverse=True)
    for i,client in enumerate(sorted_c):
        row=i+3; alt=LGR if i%2==0 else WHITE
        s=client['sales']; vals=[s.get(m,0) for m in months]
        avg=sum(vals)/4; jan=vals[3]
        prev=vals[2] if vals[2]>0 else (vals[1] if vals[1]>0 else vals[0])
        trend=(jan-prev)/prev if prev>0 else 0
        if prev==0: df,dc,arrow=FL_F,FL_C,"► FLAT"
        elif jan>prev: df,dc,arrow=UP_F,UP_C,"▲ UP"
        else: df,dc,arrow=DN_F,DN_C,"▼ DOWN"

        def sc2(col,val,bg,bold=False,color="000000",fmt=None,ha='left'):
            c=ws2.cell(row=row,column=col,value=val)
            c.font=Font(name="Arial",size=10,bold=bold,color=color)
            c.fill=PatternFill("solid",start_color=bg)
            c.alignment=Alignment(horizontal=ha,vertical='center'); c.border=bdr
            if fmt: c.number_format=fmt

        sc2(1,i+1,alt,ha='center'); sc2(2,client['name'],alt); sc2(3,client['code'],alt,ha='center')
        sc2(4,avg,LB,fmt=money,ha='right'); sc2(5,jan,LG,fmt=money,ha='right')
        sc2(6,trend,df,bold=True,color=dc,fmt='+0.0%;-0.0%;0.0%',ha='center')
        sc2(7,arrow,df,bold=True,color=dc,ha='center')
        ws2.row_dimensions[row].height=18
    ws2.freeze_panes="D3"
    return wb

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=False)
